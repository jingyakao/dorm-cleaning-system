# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import holidays

# --- 1. 網頁配置與標題 ---
st.set_page_config(page_title="智慧宿舍排程器", layout="wide")

st.title("🏨 智慧宿舍排程器")
st.write("### 資管專業實務專案：雲端智慧排程解決方案")

st.info("""
**系統亮點：**
1. **自動感知**：整合 `holidays` 演算法，自動識別台灣國定假日。
2. **動態調度**：連假週自動切換為「一二日」模式（收假清理），平常週為「二四日」。
3. **異常處理**：支援管理員手動輸入「校內特定假期」。
""")

# --- 2. 側邊欄設定 (使用者輸入介面) ---
with st.sidebar:
    st.header("⚙️ 系統參數設定")
    
    # 房號規則設定
    area_prefix = st.text_input("區域與樓層 (例如 A2)", value="A2")
    room_numbers = st.text_input("房間編號 (用逗號隔開)", value="1, 2, 3, 4, 5, 6")
    rooms = [f"{area_prefix}{n.strip()}" for n in room_numbers.split(",") if n.strip()]
    
    st.divider()
    
    # 排程基礎設定
    start_date = st.date_input("學期起始日 (週二)", value=datetime(2026, 2, 24))
    weeks_to_generate = st.number_input("生成週數", min_value=1, max_value=22, value=18)
    
    st.divider()
    
    # 手動例外日期設定
    st.subheader("🗓️ 例外假期設定")
    st.write("系統已啟動【自動國定假日判定】")
    extra_holiday_input = st.text_area(
        "新增校內假期 (格式: YYYY-MM-DD)", 
        placeholder="例如: 2026-03-15",
        help="若有校慶或學校特有放假日，請輸入於此，多筆請用逗號隔開。"
    )
    
    # 建立台灣假日庫
    tw_holidays = holidays.Taiwan(years=[start_date.year, start_date.year + 1])
    
    # 解析手動輸入日期
    try:
        extra_holidays = [datetime.strptime(d.strip(), '%Y-%m-%d').date() for d in extra_holiday_input.split(",") if d.strip()]
    except:
        st.error("日期格式錯誤，請確保為 YYYY-MM-DD")
        extra_holidays = []

# --- 3. 核心邏輯：判斷週與生成資料 ---
def is_holiday_week(monday_dt, tw_hols, ex_hols):
    # 檢查該週（一到日）是否包含任何假期
    for i in range(7):
        check_date = monday_dt + timedelta(days=i)
        if check_date in tw_hols or check_date in ex_hols:
            return True
    return False

def generate_excel(rooms, start_date, weeks, tw_hols, ex_hols):
    inspection_items = [
        "1、廁所地面及馬桶清洗乾淨無污垢及臭味", "2、廁所裡無殘留衛生紙垃圾",
        "3、浴室外地面無黃垢及排水孔無毛髮垃圾", "4、浴室內牆、門板、地面無黃垢無毛髮垃圾",
        "5、鏡子洗手台乾淨無黃垢、窗台擦拭乾淨", "6、浴室排水溝槽刷洗乾淨、無毛髮垃圾",
        "7、清掃工具及清潔用品集中擺放整齊", "8、脫水機、飲水機本身及四周地面清潔",
        "9、走道清掃並拖拭乾淨，交誼廳桌椅排整齊"
    ]

    all_rows = []
    # 確保從起始日那一週的週一開始計算
    current_monday = start_date - timedelta(days=start_date.weekday())

    for i in range(weeks):
        assigned_room = rooms[i % len(rooms)] if rooms else "N/A"
        is_holiday = is_holiday_week(current_monday, tw_hols, ex_hols)
        
        # 調度邏輯：連假週(一二日) vs 平常週(二四日)
        # offset: 0=Mon, 1=Tue, 3=Thu, 6=Sun
        target_days = [0, 1, 6] if is_holiday else [1, 3, 6]
        
        for offset in target_days:
            target_dt = current_monday + timedelta(days=offset)
            row = {
                "寢室": assigned_room,
                "打掃日期": target_dt.strftime('%Y-%m-%d (%a)'),
                "週次": f"第 {i+1:02d} 週" + (" (調整週)" if is_holiday else ""),
                "負責同學1": "", "負責同學2": "", "負責同學3": "", "負責同學4": ""
            }
            for item in inspection_items: row[item] = ""
            all_rows.append(row)
        current_monday += timedelta(days=7)

    df = pd.DataFrame(all_rows)
    
    # --- 報表美化 ---
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='打掃分配表')
        ws = writer.sheets['打掃分配表']
        
        # 樣式定義
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)
        center_ali = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # 標題樣式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = font_white
            cell.alignment = center_ali
            cell.border = thin_border

        # 內容樣式與自動調寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell.alignment = center_ali
                cell.border = thin_border
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 5, 40)

    output.seek(0)
    return output

# --- 4. 主畫面顯示與下載功能 ---
if not rooms:
    st.warning("👈 請先在側邊欄輸入房間編號。")
else:
    # 預先生成檔案
    excel_data = generate_excel(rooms, start_date, weeks_to_generate, tw_holidays, extra_holidays)

    # 下載區塊
    col1, col2 = st.columns([1, 3])
    with col1:
        st.download_button(
            label="📥 下載 Excel 分配表",
            data=excel_data,
            file_name=f"{area_prefix}_宿舍排程_{datetime.now().strftime('%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col2:
        st.success(f"已成功規劃 {weeks_to_generate} 週排程，共 {len(rooms)} 間寢室輪替。")

    # 資料預覽
    st.divider()
    st.subheader("👀 排程內容預覽")
    excel_data.seek(0)
    df_preview = pd.read_excel(excel_data)
    st.dataframe(df_preview, use_container_width=True)
